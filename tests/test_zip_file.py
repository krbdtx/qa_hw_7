import os.path
import pypdf
from zipfile import ZipFile
from openpyxl import load_workbook

parent_dir = os.path.dirname(os.path.dirname(__file__))
source_dir = os.path.join(parent_dir, 'resources')
archive = os.path.join(source_dir, 'zip_file.zip')


def test_data_in_pdf():
    with ZipFile(archive, 'r') as zip_arch:
        with zip_arch.open("file_pdf.pdf") as file_source:
            readf = pypdf.PdfReader(file_source)
            page = (readf.pages[4].extract_text())
    assert "cppdf .msi" in page


def test_data_in_xlsx():
    with ZipFile(archive, 'r') as zip_arch:
        with zip_arch.open("file_xlsx.xlsx") as file_source:
            readf = load_workbook(file_source)
            sheet = readf.active
            count = (sheet.cell(6, 3).value)
    assert count == "Техническая поддержка"


def test_data_in_csv():
    with ZipFile(archive, 'r') as zip_arch:
        with zip_arch.open('file_csv.csv') as file_source:
            readf = file_source.read().decode('utf-8')
    assert "MUST SELL" in readf
